import os
from os import sep
from tkinter.constants import END
from typing import Text
import pandas as pd
import tkinter as tk
from tkinter import Button, Entry, Label, Message, Pack, Toplevel, filedialog
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
import subprocess
from datetime import date, datetime

hsm=tk.Tk()

hsm.title("시행기록지 메이커")
hsm.geometry("640x300+100+100")
hsm.resizable(True,True)

fileA=""
fileB=""
fileAA=Label(hsm,text=fileA)
#fileBB=Label(hsm,text=fileB)
result=Label(hsm,text=" ")

def Load1():
    filename = filedialog.askopenfilename(initialdir=r"C:\Users\tmdal\Desktop", title="Select file",
                                          filetypes=(("xlsx files", "*.xlsx"),
                                          ("xlsx files","*.xls"),
                                          ("all files", "*.*")))
    fileAA.configure(text="이온엠파일: " + filename)
    global fileA
    fileA=filename


fileAA.pack()

'''def Load2():
    filename = filedialog.askopenfilename(initialdir=r"C:\재활스케쥴\2021년", title="Select file",
                                          filetypes=(("xlsx files", "*.xlsx"),
                                          ("xlsx files","*.xls"),
                                          ("all files", "*.*")))
    fileBB.configure(text="시행기록지: " + filename)
    global fileB
    fileB=filename

fileBB.pack()'''

def closemenu():
    hsm.quit()
    hsm.destroy()

def activate():
    activelabel=Label(text=eonm_to_timetable())
    activelabel.pack()

def activate2():
    activelabel2=Label(text=eonm_to_timetable2())
    activelabel2.pack()
    
def indexerror():
    tk.messagebox.showerror(title=indexerror, message="인덱스 에러!!\n 해결방법: 1. 시행기록지base 파일내 누락된 이름을 확인하세요.\n2. 이온엠 파일이 '시행완료'인지 확인하세요.")

def basefileerror():
    tk.messagebox.showerror(title="파일이 없습니다.", message="실행폴더 내에 시행기록지base 파일이 없거나 이름이 변경되었습니다 확인하세요.")

'''def pandasxl():
    openwbA= pd.read_excel(fileA,"Sheet1")
    openwbB= pd.read_excel(fileB,"물리 출력용")
    PTnameA=list(openwbA['치료사'])
    PTnameB=list(openwbB.iloc[2,:])
    sumnumber=list(openwbB.iloc[71,:])
    finalwin=Toplevel(hsm)
    def namecount(a):
        return(PTnameA.count(a))
        
    for i in range(int(len(PTnameB)/2+1)):
        if namecount(PTnameB[(i)*2-1]) == sumnumber[(i)*2]:
            resultok=Label(finalwin,text="")
            resultok.pack()
        else:
            resultwrong=Label(finalwin,text=PTnameB[(i)*2-1])
            resultwrong.pack()
'''
def eonm_to_timetable():
    try:
        openwbA= pd.read_excel(fileA,"Sheet1",usecols="D,J,M,T")
        openwbB= pd.read_excel(os.getcwd()+"\시행기록지 base.xlsx","물치")
    except FileNotFoundError:
        basefileerror()
    
    thincol=['b','d','f','h','j','l','n','p','r','t','v','x','z','ab','ad','af','ah','aj','al','an','ap','ar','at','av','ax','az']
    thickcol=['c','e','g','i','k','m','o','q','s','u','w','y','aa','ac','ae','ag','ai','ak','am','ao','aq','as','au','aw','ay','ba',]


    #try:
    for i in range(len(openwbA)):
        a=openwbB.where(openwbB==openwbA.iloc[i,2]).dropna(how='all').dropna(axis=1)
        b=openwbB.where(openwbB==openwbA.iloc[i,1][0:5]).dropna(how='all').dropna(axis=1)

        try:
            if pd.isna(openwbB.iloc[b.index[0],a.columns[0]]):
                openwbB.iloc[b.index[0],a.columns[0]]=openwbA.iloc[i,3]
                openwbB.iloc[b.index[0],a.columns[0]+1]=openwbA.iloc[i,0]
            elif pd.isna(openwbB.iloc[b.index[0]+1,a.columns[0]]):
                openwbB.iloc[b.index[0]+1,a.columns[0]]=openwbA.iloc[i,3]
                openwbB.iloc[b.index[0]+1,a.columns[0]+1]=openwbA.iloc[i,0]
            elif pd.isna(openwbB.iloc[b.index[0]+2,a.columns[0]]):
                openwbB.iloc[b.index[0]+2,a.columns[0]]=openwbA.iloc[i,3]
                openwbB.iloc[b.index[0]+2,a.columns[0]+1]=openwbA.iloc[i,0]
            elif pd.isna(openwbB.iloc[b.index[0]+3,a.columns[0]]):
                openwbB.iloc[b.index[0]+3,a.columns[0]]=openwbA.iloc[i,3]
                openwbB.iloc[b.index[0]+3,a.columns[0]+1]=openwbA.iloc[i,0]
            elif len(openwbB.iloc[b.index[0],a.columns[0]+1])<5:          ##
                openwbB.iloc[b.index[0],a.columns[0]+1]=openwbB.iloc[b.index[0],a.columns[0]+1]+"/"+openwbA.iloc[i,0]
            elif len(openwbB.iloc[b.index[0]+1,a.columns[0]+1])<5:
                openwbB.iloc[b.index[0]+1,a.columns[0]+1]=openwbB.iloc[b.index[0]+1,a.columns[0]+1]+"/"+openwbA.iloc[i,0]
            elif len(openwbB.iloc[b.index[0]+2,a.columns[0]+1])<5:
                openwbB.iloc[b.index[0]+2,a.columns[0]+1]=openwbB.iloc[b.index[0]+2,a.columns[0]+1]+"/"+openwbA.iloc[i,0]
            elif len(openwbB.iloc[b.index[0]+3,a.columns[0]+1])<5:
                openwbB.iloc[b.index[0]+3,a.columns[0]+1]=openwbB.iloc[b.index[0]+3,a.columns[0]+1]+"/"+openwbA.iloc[i,0]
            elif len(openwbB.iloc[b.index[0],a.columns[0]+1])<9:          ##
                openwbB.iloc[b.index[0],a.columns[0]+1]=openwbB.iloc[b.index[0],a.columns[0]+1]+"/"+openwbA.iloc[i,0]
            elif len(openwbB.iloc[b.index[0]+1,a.columns[0]+1])<9:
                openwbB.iloc[b.index[0]+1,a.columns[0]+1]=openwbB.iloc[b.index[0]+1,a.columns[0]+1]+"/"+openwbA.iloc[i,0]
            elif len(openwbB.iloc[b.index[0]+2,a.columns[0]+1])<9:
                openwbB.iloc[b.index[0]+2,a.columns[0]+1]=openwbB.iloc[b.index[0]+2,a.columns[0]+1]+"/"+openwbA.iloc[i,0]
            elif len(openwbB.iloc[b.index[0]+3,a.columns[0]+1])<9:
                openwbB.iloc[b.index[0]+3,a.columns[0]+1]=openwbB.iloc[b.index[0]+3,a.columns[0]+1]+"/"+openwbA.iloc[i,0]
        except:
            print(i)
    openwbB.iloc[1,23]="프라임재활센터"
    openwbB.iloc[1,41]="프라임재활센터"
    datexl=os.path.basename(fileA)
    openwbB.iloc[1,27]=datetime.strptime(datexl[0:4]+"-"+datexl[4:6]+"-"+datexl[6:8],'%Y-%m-%d')
    openwbB.iloc[1,45]=datetime.strptime(datexl[0:4]+"-"+datexl[4:6]+"-"+datexl[6:8],'%Y-%m-%d')
    openwbB.replace("중추신경계발달재활치료 ","N",inplace=True)
    openwbB.replace("기능적전기자극치료(FES) ","F",inplace=True)
    openwbB.replace("재활기능치료-매트 및 이동치료 ","M",inplace=True)
    openwbB.replace("재활기능치료-보행치료 ","G",inplace=True)
    openwbB.replace("R -TMS 3 ","T",inplace=True)
    openwbB.replace('관절가동범위검사 ',"A",inplace=True)
    openwbB.replace('신경계도수2.5 ',"도h",inplace=True)
    openwbB.replace('도수근력검사(전신) ',"A",inplace=True)
    openwbB.replace('신경계도수치료(Manual Therapy)6 ',"도",inplace=True)
    openwbB.replace('신경계도수치료(Manual Therapy)5 ',"도",inplace=True)

    openwbB.to_excel(os.getcwd()+'\hsmprogram.xlsx')


    ##########################
    makeshape=openpyxl.load_workbook(os.getcwd()+'\hsmprogram.xlsx')
    makeshapews=makeshape["Sheet1"]
    makeshapews.delete_cols(1)
    makeshapews.delete_rows(1)
    greanic_color=PatternFill(start_color='b3ff99',end_color='b3ff99',fill_type='solid')
    def fontcontrol(a):
        return Font(size=a)
    #number_format='yyyy"년"mm"월"dd"일"aaa'

    #셀 크기 조정

    for row in range(4,60):
        makeshapews.row_dimensions[row].height=16.5
    makeshapews.row_dimensions[2].height=40
    makeshapews.column_dimensions['A'].width=19
    for name in thincol:
        makeshapews.column_dimensions[name].width=4
    for name in thickcol:
        makeshapews.column_dimensions[name].width=15
    for i in range(1,27):
        makeshapews.merge_cells(start_row=3,start_column=i*2,end_row=3,end_column=i*2+1)
        makeshapews.cell(3,i*2).fill=greanic_color
        makeshapews.cell(3,i*2).alignment=Alignment(horizontal='center')
        makeshapews.cell(3,i*2).font=fontcontrol(20)
        for o in range(1,61):
            makeshapews.cell(o+1,i*2).border=Border(left=Side(style='thin'))
            makeshapews.cell(o+1,1).border=Border(left=Side(style='thin'))

    for i in range (1,52):
        makeshapews.cell(2,i).border=Border(top=Side(style='thin'))
        makeshapews.cell(3,i).border=Border(top=Side(style='thin'))
        makeshapews.cell(62,i).border=Border(top=Side(style='thin'))
        for o in range(1,9):
            makeshapews.cell(o*4,i).border=Border(top=Side(style='thin'))
            makeshapews.cell(o*4+29,i).border=Border(top=Side(style='thin'))
    for i in range(1,26):
        for o in range (1,8):
            makeshapews.cell(o*4,1).border=Border(top=Side(style='thin'),left=Side(style='thin'))
            makeshapews.cell(o*4+29,1).border=Border(top=Side(style='thin'),left=Side(style='thin'))
            makeshapews.cell(3,i*2).border=Border(top=Side(style='thin'),left=Side(style='thin'))
            makeshapews.cell(o*4,i*2).border=Border(top=Side(style='thin'),left=Side(style='thin'))
            makeshapews.cell(o*4+29,i*2).border=Border(top=Side(style='thin'),left=Side(style='thin'))
    makeshapews.cell(2,1).border=Border(top=Side(style='thin'),left=Side(style='thin'))
    makeshapews.cell(3,1).border=Border(top=Side(style='thin'),left=Side(style='thin'))
    makeshapews.cell(32,1).border=Border(top=Side(style='thin'),left=Side(style='thin'))
    makeshapews.cell(61,1).border=Border(top=Side(style='thin'),left=Side(style='thin'))
    makeshapews.merge_cells(start_row=2,start_column=24,end_row=2,end_column=27)
    makeshapews.merge_cells(start_row=2,start_column=28,end_row=2,end_column=31)
    makeshapews.merge_cells(start_row=2,start_column=42,end_row=2,end_column=45)
    makeshapews.merge_cells(start_row=2,start_column=46,end_row=2,end_column=49)
    for i in range(1,15):
        makeshapews.merge_cells(start_row=i*2+2,start_column=1,end_row=i*2+3,end_column=1)
        makeshapews.merge_cells(start_row=i*2+31,start_column=1,end_row=i*2+32,end_column=1)
        makeshapews.cell(i*2+2,1).alignment=Alignment(horizontal='center')
        makeshapews.cell(i*2+2,1).font=fontcontrol(20)
        makeshapews.cell(i*2+31,1).font=fontcontrol(20)
        makeshapews.cell(i*2+31,1).alignment=Alignment(horizontal='center')
    for i in range(4,60):
        for o in range(2,52):
            makeshapews.cell(i,o).font=fontcontrol(12)

    makeshapews.cell(2,28).number_format='yyyy"년"mm"월"dd"일"aaa"요일"'
    makeshapews.cell(2,28).font=fontcontrol(20)
    makeshapews.cell(2,46).number_format='yyyy"년"mm"월"dd"일"aaa"요일"'
    makeshapews.cell(2,46).font=fontcontrol(20)
    makeshapews.cell(2,24).font=fontcontrol(20)
    makeshapews.cell(2,42).font=fontcontrol(20)
    '''for col in range(1,makeshapews.max_column):
            makeshapews.column_dimensions[chr(col)].width=4'''
        #print area
        #makeshapews.print_area = 'A2:AG61'
    #except IndexError:
        #indexerror()
     #   print('fuck')
    
    
    makeshape.save(os.getcwd()+'\hsmprogram.xlsx')
    subprocess.run(["C:\Program Files (x86)\Hnc\Office 2020\HOffice110\Bin\HCell.exe", os.getcwd()+'\hsmprogram.xlsx'])


def eonm_to_timetable2():
    try:
        openwbA= pd.read_excel(fileA,"Sheet1",usecols="D,J,M,T")
        openwbB= pd.read_excel(os.getcwd()+"\시행기록지 base.xlsx","작치")
    except FileNotFoundError:
        basefileerror()
    
    thincol=['b','d','f','h','j','l','n','p','r','t','v','x','z','ab','ad','af','ah','aj','al','an','ap','ar','at','av','ax','az']
    thickcol=['c','e','g','i','k','m','o','q','s','u','w','y','aa','ac','ae','ag','ai','ak','am','ao','aq','as','au','aw','ay','ba',]

    vital=0
    plusone=int(1)

    #try:
    for i in range(len(openwbA)):
        a=openwbB.where(openwbB==openwbA.iloc[i,2]).dropna(how='all').dropna(axis=1)
        b=openwbB.where(openwbB==openwbA.iloc[i,1]).dropna(how='all').dropna(axis=1)
        openwbB.iloc[60,a.columns[0]+1]=()
        
        try:
            if pd.isna(openwbB.iloc[b.index[0],a.columns[0]]):
                openwbB.iloc[b.index[0],a.columns[0]]=openwbA.iloc[i,3]
                openwbB.iloc[b.index[0],a.columns[0]+1]=openwbA.iloc[i,0]
            elif pd.isna(openwbB.iloc[b.index[0]+1,a.columns[0]]):
                openwbB.iloc[b.index[0]+1,a.columns[0]]=openwbA.iloc[i,3]
                openwbB.iloc[b.index[0]+1,a.columns[0]+1]=openwbA.iloc[i,0]
            elif pd.isna(openwbB.iloc[b.index[0]+2,a.columns[0]]):
                openwbB.iloc[b.index[0]+2,a.columns[0]]=openwbA.iloc[i,3]
                openwbB.iloc[b.index[0]+2,a.columns[0]+1]=openwbA.iloc[i,0]
            elif pd.isna(openwbB.iloc[b.index[0]+3,a.columns[0]]):
                openwbB.iloc[b.index[0]+3,a.columns[0]]=openwbA.iloc[i,3]
                openwbB.iloc[b.index[0]+3,a.columns[0]+1]=openwbA.iloc[i,0]
        except:
            1
        '''if openwbA.iloc[i,3] == "연하재활 기능적전기자극치료 ":
            openwbB.iloc[60,a.columns[0]+1]=openwbB.iloc[60,a.columns[0]+1]+plusone ##########################'''
        
    openwbB.iloc[1,23]="프라임재활센터"
    openwbB.iloc[1,41]="프라임재활센터"
    datexl=os.path.basename(fileA)
    openwbB.iloc[1,27]=datetime.strptime(datexl[0:4]+"-"+datexl[4:6]+"-"+datexl[6:8],'%Y-%m-%d')
    #openwbB.iloc[1,45]=datetime.strptime(datexl[0:4]+"-"+datexl[4:6]+"-"+datexl[6:8],'%Y-%m-%d')
    openwbB.replace("연하장애재활치료 ","연",inplace=True)
    openwbB.replace("작업치료-특수작업치료 ","S",inplace=True)
    openwbB.replace("일상생활동작 훈련치료 ADL[1일당] ","A",inplace=True)
    openwbB.replace("작업치료-복합작업치료 ","C",inplace=True)
    openwbB.replace("전산화 인지재활치료(30분) ","인",inplace=True)
    openwbB.replace('수지기능검사(젭슨수부평가검사) ',"평j",inplace=True)
    openwbB.replace('일상생활동작검사-SCIM평가 ',"평s",inplace=True)
    openwbB.replace('일상생활동작검사- MBI(변형된 바델지수등을 이용한 경우) ',"평m",inplace=True)

    openwbB.to_excel(os.getcwd()+'\hsmprogram.xlsx')


    ##########################
    makeshape=openpyxl.load_workbook(os.getcwd()+'\hsmprogram.xlsx')
    makeshapews=makeshape["Sheet1"]
    makeshapews.delete_cols(1)
    makeshapews.delete_rows(1)
    greanic_color=PatternFill(start_color='b3ff99',end_color='b3ff99',fill_type='solid')
    def fontcontrol(a):
        return Font(size=a)
    #number_format='yyyy"년"mm"월"dd"일"aaa'

    #셀 크기 조정

    for row in range(4,60):
        makeshapews.row_dimensions[row].height=16.5
    makeshapews.row_dimensions[2].height=40
    makeshapews.column_dimensions['A'].width=19
    for name in thincol:
        makeshapews.column_dimensions[name].width=4
    for name in thickcol:
        makeshapews.column_dimensions[name].width=15
    for i in range(1,27):
        makeshapews.merge_cells(start_row=3,start_column=i*2,end_row=3,end_column=i*2+1)
        makeshapews.cell(3,i*2).fill=greanic_color
        makeshapews.cell(3,i*2).alignment=Alignment(horizontal='center')
        makeshapews.cell(3,i*2).font=fontcontrol(20)
        for o in range(1,61):
            makeshapews.cell(o+1,i*2).border=Border(left=Side(style='thin'))
            makeshapews.cell(o+1,1).border=Border(left=Side(style='thin'))

    for i in range (1,52):
        makeshapews.cell(2,i).border=Border(top=Side(style='thin'))
        makeshapews.cell(3,i).border=Border(top=Side(style='thin'))
        makeshapews.cell(62,i).border=Border(top=Side(style='thin'))
        for o in range(1,9):
            makeshapews.cell(o*4,i).border=Border(top=Side(style='thin'))
            makeshapews.cell(o*4+29,i).border=Border(top=Side(style='thin'))
    for i in range(1,26):
        for o in range (1,8):
            makeshapews.cell(o*4,1).border=Border(top=Side(style='thin'),left=Side(style='thin'))
            makeshapews.cell(o*4+29,1).border=Border(top=Side(style='thin'),left=Side(style='thin'))
            makeshapews.cell(3,i*2).border=Border(top=Side(style='thin'),left=Side(style='thin'))
            makeshapews.cell(o*4,i*2).border=Border(top=Side(style='thin'),left=Side(style='thin'))
            makeshapews.cell(o*4+29,i*2).border=Border(top=Side(style='thin'),left=Side(style='thin'))
    makeshapews.cell(2,1).border=Border(top=Side(style='thin'),left=Side(style='thin'))
    makeshapews.cell(3,1).border=Border(top=Side(style='thin'),left=Side(style='thin'))
    makeshapews.cell(32,1).border=Border(top=Side(style='thin'),left=Side(style='thin'))
    makeshapews.cell(61,1).border=Border(top=Side(style='thin'),left=Side(style='thin'))
    makeshapews.merge_cells(start_row=2,start_column=24,end_row=2,end_column=27)
    makeshapews.merge_cells(start_row=2,start_column=28,end_row=2,end_column=31)
    makeshapews.merge_cells(start_row=2,start_column=42,end_row=2,end_column=45)
    makeshapews.merge_cells(start_row=2,start_column=46,end_row=2,end_column=49)
    for i in range(1,15):
        makeshapews.merge_cells(start_row=i*2+2,start_column=1,end_row=i*2+3,end_column=1)
        makeshapews.merge_cells(start_row=i*2+31,start_column=1,end_row=i*2+32,end_column=1)
        makeshapews.cell(i*2+2,1).alignment=Alignment(horizontal='center')
        makeshapews.cell(i*2+2,1).font=fontcontrol(20)
        makeshapews.cell(i*2+31,1).font=fontcontrol(20)
        makeshapews.cell(i*2+31,1).alignment=Alignment(horizontal='center')
    for i in range(4,60):
        for o in range(2,52):
            makeshapews.cell(i,o).font=fontcontrol(12)

    makeshapews.cell(2,28).number_format='yyyy"년"mm"월"dd"일"aaa"요일"'
    makeshapews.cell(2,28).font=fontcontrol(20)
    makeshapews.cell(2,46).number_format='yyyy"년"mm"월"dd"일"aaa"요일"'
    makeshapews.cell(2,46).font=fontcontrol(20)
    makeshapews.cell(2,24).font=fontcontrol(20)
    makeshapews.cell(2,42).font=fontcontrol(20)
    '''for col in range(1,makeshapews.max_column):
            makeshapews.column_dimensions[chr(col)].width=4'''
        #print area
        #makeshapews.print_area = 'A2:AG61'
    #except IndexError:
        #indexerror()
     #   print('fuck')
    
    
    makeshape.save(os.getcwd()+'\hsmprogram.xlsx')
    subprocess.run(["C:\Program Files (x86)\Hnc\Office 2020\HOffice110\Bin\HCell.exe", os.getcwd()+'\hsmprogram.xlsx'])



menubar=tk.Menu(hsm)

menu_1=tk.Menu(menubar,tearoff=0)

menu_1.add_command(label="이온엠 파일",command=Load1)
#menu_1.add_command(label="시행기록지",command=Load2)
#menu_1.add_command(label="지우기")
menu_1.add_separator()
menu_1.add_command(label="닫기",command=closemenu)
menubar.add_cascade(label="파일",menu=menu_1)

#menu_2=tk.Menu(menubar,tearoff=0,selectcolor="red")
#menu_2.add_command(label="이건아직")
#menu_2.add_command(label="안만들어짐")
#menu_2.add_command(label="패치를")
#menu_2.add_command(label="기다리세요")
#menubar.add_cascade(label="뭐지 이건",menu=menu_2)
hsm.config(menu=menubar)

activate_button=Button(hsm,text="PT",command=activate)
activate_button.pack()


activate_button2=Button(hsm,text="OT",command=activate2)
activate_button2.pack()

    
worning=Label(hsm,text="주의: 1.파일을 저장하세요 안그럼 다시하셔야할껍니다.\n 2.하나의 파일을 변환한 뒤 꼭 닫은 후 다음 파일을 진행하세요(어차피 안될껍니다)")
worning.pack()
wewill=Label(hsm,text="추후 추가사항: 1.아직 프린트 범위는 적용시키지 못했습니다 현재 시도중입니다.")
wewill.pack()
waiting=Label(hsm,text='파일을 변환하는데 대략 20초정도 걸립니다.\n 인내심을 가지고 기다리세요')
waiting.pack()

hsm.mainloop()
